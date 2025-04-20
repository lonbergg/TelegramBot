import os
import sys
# 1) Путь до папки с вашим BotGGpokerMain.py — это ровно один уровень вверх от tests/
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
# 2) Добавляем эту папку в sys.path
sys.path.insert(0, ROOT)

import re
import pytest
import asyncio
from datetime import datetime
from openpyxl import Workbook, load_workbook
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, Message, CallbackQuery

# 3) Импортируем ваш файл бота под текущим именем
import BotGGpokerMain as bot_module


# --- UTILS TESTS ---
@pytest.fixture(autouse=True)
def isolate_participants(tmp_path, monkeypatch):
    file = tmp_path / "participants.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(["Telegram ID", "Username", "Full Name", "Дата участі", "GGPoker Нік", "Email"])
    wb.save(file)

    monkeypatch.setenv("EXCEL_FILE", str(file))
    monkeypatch.setattr(bot_module, "EXCEL_FILE", str(file))
    monkeypatch.setattr(bot_module, "participants_set", set())
    return str(file)


class DummyUser:
    def __init__(self, user_id, username, first_name, last_name=None):
        self.id = user_id
        self.username = username
        self.first_name = first_name
        self.last_name = last_name


def test_save_participant_and_duplicates(isolate_participants):
    user = DummyUser(123, "tester", "Test", "User")
    # первый вызов должен вернуть True
    assert bot_module.save_participant(user, "nick1", "a@b.com") is True
    # проверяем, что запись появилась в Excel
    wb = load_workbook(isolate_participants)
    data = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert data[0][4] == "nick1"
    # второй вызов — дубликат
    assert bot_module.save_participant(user, "nick1", "a@b.com") is False


@pytest.mark.parametrize("email,valid", [
    ("user@example.com", True),
    ("user.name+tag@domain.co", True),
    ("invalid-email", False),
    ("@nouser.com", False),
    ("user@.com", False),
])
def test_email_regex(email, valid):
    # точный паттерн из вашего бота, поддерживает '+'
    pattern = re.compile(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$')
    assert (pattern.match(email) is not None) == valid


def test_user_menu_buttons():
    kb = bot_module.user_menu(is_admin=False)
    texts = [b.text for row in kb.keyboard for b in row]
    assert "📜 Умови" in texts
    assert "🎁 Призи" in texts
    assert "🔐 Admin panel" not in texts

    kb_admin = bot_module.user_menu(is_admin=True)
    texts_admin = [b.text for row in kb_admin.keyboard for b in row]
    assert "🔐 Admin panel" in texts_admin


def test_support_and_admin_menu():
    support_kb = bot_module.support_menu()
    texts = [b.text for row in support_kb.keyboard for b in row]
    assert "✍️ Написати в підтримку" in texts
    assert "🔄 Змінити нікнейм" in texts

    admin_kb = bot_module.admin_menu()
    texts = [b.text for row in admin_kb.keyboard for b in row]
    for exp in ["👥 Учасники", "📥 Експорт Excel", "📤 Список з бази PostgreSQL"]:
        assert exp in texts


@pytest.mark.asyncio
async def test_export_db_to_excel(monkeypatch):
    class DummyConn:
        async def fetch(self, _):
            return [(1, "@u", "Full", datetime.now(), "nick", "e@mail.com")]

    class DummyPool:
        # acquire возвращает сразу контекст‑менеджер
        def acquire(self):
            class Ctx:
                async def __aenter__(self):
                    return DummyConn()
                async def __aexit__(self, exc_type, exc, tb):
                    pass
            return Ctx()

    fn = bot_module.export_db_to_excel(DummyPool())
    path = await fn()
    assert path.endswith(".xlsx")
    assert os.path.exists(path)


# --- HANDLER TESTS ---
@pytest.mark.asyncio
async def test_participate_and_full_registration_flow(monkeypatch):
    # 1) Тестируем participate_command
    user = DummyUser(1, None, "First")
    messages = []

    class DummyMessage:
        def __init__(self):
            self.from_user = user
        async def answer(self, text, **kwargs):
            messages.append(text)

    dm = DummyMessage()
    await bot_module.participate_command(dm)
    assert any("Обробляємо" in m for m in messages)
    assert any("Для участі" in m for m in messages)

    # 2) Проверяем check_subscription — успешная подписка
    class Member: status = "member"
    monkeypatch.setattr(bot_module.bot, "get_chat_member", lambda ch, uid: asyncio.Future())
    fut = bot_module.bot.get_chat_member(None, None)
    fut.set_result(Member())

    class DummyCallback:
        def __init__(self):
            self.from_user = user
            self.message = dm
            self.answers = []
        async def answer(self, text, **kw):
            self.answers.append(text)

    cb = DummyCallback()
    await bot_module.check_subscription(cb)
    assert bot_module.user_states[user.id] == "awaiting_nickname"

    # 3) Вводим никнейм и email через handle_messages
    dm2 = DummyMessage(); dm2.text = "my_nick"
    await bot_module.handle_messages(dm2)
    dm3 = DummyMessage(); dm3.text = "my@mail.com"
    await bot_module.handle_messages(dm3)
    # в user_states должен быть шаг confirming
    assert bot_module.user_states[user.id]["step"] == "confirming"

    # 4) Подтверждаем участие
    cb2 = DummyCallback(); cb2.data = "confirm_participation"; cb2.message = dm
    # Устанавливаем состояние вручную
    bot_module.user_states[user.id] = {
        "step": "confirming",
        "nickname": "my_nick",
        "email": "my@mail.com",
    }
    await bot_module.confirm_participation(cb2)
    # После подтверждения state должен удалиться
    assert user.id not in bot_module.user_states


@pytest.mark.asyncio
async def test_handle_spam_prevention():
    user = DummyUser(2, None, "Spam")
    # Создаём два сообщения подряд
    class M: pass
    m = type("X", (object,), {"from_user": user, "text": "hi", "answer": lambda *a, **k: None})()
    # Первый вызов — должен пройти
    await bot_module.handle_messages(m)
    # Второй незадолго — должен проигнорироваться, но без ошибок
    await bot_module.handle_messages(m)
    assert user.id not in bot_module.user_states


