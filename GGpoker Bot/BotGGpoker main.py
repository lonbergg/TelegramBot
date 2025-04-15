import logging
import os
import re
import asyncio
import asyncpg
import pandas as pd
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, Router, types, F
from aiogram.filters import Command
from aiogram.types import (
    InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
    Message, CallbackQuery, FSInputFile
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.date import DateTrigger
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

API_TOKEN = os.getenv("API_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))
DATABASE_HOST = os.getenv("DATABASE_HOST")
DATABASE_PORT = int(os.getenv("DATABASE_PORT"))
DATABASE_USER = os.getenv("DATABASE_USER")
DATABASE_PASSWORD = os.getenv("DATABASE_PASSWORD")
DATABASE_NAME = os.getenv("DATABASE_NAME")
CHANNEL_USERNAME = os.getenv("CHANNEL_USERNAME")
YOUTUBE_LINK = os.getenv("YOUTUBE_LINK")
TWITCH_LINK = os.getenv("TWITCH_LINK")
SUPPORT_USERNAME = os.getenv("SUPPORT_USERNAME")

bot = Bot(token=API_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler()
router = Router()
logging.basicConfig(level=logging.INFO)

EXCEL_FILE = 'participants.xlsx'
LOG_FILE = 'broadcast_log.xlsx'
user_states = {}
participants_set = set()
admin_states = {}
banned_users = set()
last_message_times = {}
broadcast_buffer = {}

SPAM_INTERVAL = timedelta(seconds=5)

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(["Telegram ID", "Username", "Full Name", "Дата участі", "GGPoker Нік", "Email"])
    wb.save(EXCEL_FILE)

if not os.path.exists(LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.append(["Дата", "Повідомлення", "Успішно", "Не вдалося"])
    wb.save(LOG_FILE)

def export_db_to_excel(pool):
    async def inner():
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM participants")
            df = pd.DataFrame(rows, columns=["telegram_id", "username", "full_name", "joined_at", "nickname", "email"])
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"participants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            df.to_excel(filename, index=False)
            return filename
    return inner

def save_participant(user: types.User, nickname: str, email: str):
    if user.id in participants_set:
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if nickname == row[4] or user.id == row[0]:
            return False
    ws.append([
        user.id,
        f"@{user.username}" if user.username else "(без username)",
        full_name,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        nickname,
        email
    ])
    wb.save(EXCEL_FILE)
    participants_set.add(user.id)
    return True

async def save_participant_to_db(pool, user: types.User, nickname: str, email: str):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO participants (telegram_id, username, full_name, joined_at, nickname, email)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (telegram_id) DO NOTHING
            """,
            user.id,
            f"@{user.username}" if user.username else "(без username)",
            f"{user.first_name or ''} {user.last_name or ''}".strip(),
            datetime.now(),
            nickname,
            email
        )


def user_menu(is_admin=False):
    buttons = [
        [KeyboardButton(text="📜 Умови"), KeyboardButton(text="🎁 Призи")],
        [KeyboardButton(text="📞 Підтримка"), KeyboardButton(text="📍 Мій статус")],
        [KeyboardButton(text="🎉 Взяти участь у розiграшi"), KeyboardButton(text="❓ FAQ")]
    ]
    if is_admin:
        buttons.append([KeyboardButton(text="🔐 Admin panel")])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True, input_field_placeholder="Оберіть опцію")


def admin_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👥 Учасники"), KeyboardButton(text="📥 Експорт Excel")],
            [KeyboardButton(text="📤 Список з бази PostgreSQL")],
            [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="📣 Розсилка")],
            [KeyboardButton(text="🕒 Планувати розсилку"), KeyboardButton(text="⛔ Забанені")],
            [KeyboardButton(text="↩️ Повернутись")]
        ],
        resize_keyboard=True
    )


@router.message(Command("start"))
async def welcome_user(message: Message):
    if message.from_user.id in banned_users:
        return
    await message.answer(
        "👋 Вітаємо у GGpoker Telegram боті! Це не просто бот для участі в розіграші, а також ваш персональний асистент для отримання новин, бонусів та корисної інформації про GGpoker. Натисніть кнопку нижче, щоб дізнатися більше або одразу розпочати участь у розіграші!",
        reply_markup=user_menu(message.from_user.id == ADMIN_ID)
    )


@router.message(F.text == "🎉 Взяти участь у розiграшi")
async def instructions(message: Message):
    if message.from_user.id in banned_users:
        return
    await message.answer("🔄 Обробляємо запит...", disable_notification=True)
    await asyncio.sleep(1.2)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Я підписався", callback_data="participate")]
    ])
    await message.answer(
        text=(
            "📋 Для участі в розіграші потрібно:\n"
            f"1. Підписатися на Telegram канал: {CHANNEL_USERNAME}\n"
            f"2. Підписатися на YouTube: {YOUTUBE_LINK}\n"
            f"3. Підписатися на Twitch: {TWITCH_LINK}\n\n"
            "Після цього натисніть кнопку нижче, щоб продовжити."
        ),
        reply_markup=kb
    )


@router.callback_query(F.data == "participate")
async def process_participation(callback: CallbackQuery):
    if callback.from_user.id in banned_users:
        return
    user = callback.from_user
    await callback.message.answer("🔍 Перевіряємо підписку...", disable_notification=True)
    await asyncio.sleep(1.5)
    member = await bot.get_chat_member(CHANNEL_USERNAME, user.id)
    if member.status not in ["member", "administrator", "creator"]:
        await callback.answer("❌ Спочатку підпишіться на Telegram-канал!", show_alert=True)
        return
    user_states[user.id] = 'awaiting_nickname'
    await callback.message.answer("✅ Ви приєдналися! Введіть ваш GGPoker нікнейм.")
    await callback.answer()


@router.message()
async def handle_messages(message: Message):
    user_id = message.from_user.id
    text = message.text

    if user_id in banned_users:
        return

    now = datetime.now()
    last_time = last_message_times.get(user_id)
    if last_time and now - last_time < SPAM_INTERVAL:
        return
    last_message_times[user_id] = now

    if user_id in user_states:
        state = user_states[user_id]
        if state == 'awaiting_nickname':
            user_states[user_id] = {'step': 'awaiting_email', 'nickname': text.strip()}
            await message.reply("📧 Введіть вашу електронну пошту:")
            return
        elif isinstance(state, dict) and state.get('step') == 'awaiting_email':
            nickname = state['nickname']
            email = text.strip()
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                await message.reply("❌ Невірний формат email. Спробуйте ще раз:")
                return
            if save_participant(message.from_user, nickname, email):
                await save_participant_to_db(dp['db'], message.from_user, nickname, email)
                await message.reply("✅ Ваші дані збережено!", reply_markup=user_menu(message.from_user.id == ADMIN_ID))
                user_states.pop(user_id)
            else:
                banned_users.add(user_id)
                await message.reply(
                    "🚫 Ви вже брали участь у розіграші або намагалися обманути систему.\n\n"
                    "ℹ️ Але ви все одно можете користуватись ботом!"
                )
            return

    if text == "📤 Список з бази PostgreSQL" and user_id == ADMIN_ID:
        await message.answer("🔄 Експортуємо список з PostgreSQL у файл...")
        export_func = export_db_to_excel(dp['db'])
        file_path = await export_func()
        await message.answer_document(FSInputFile(file_path))
        return

    if text == "📜 Умови":
        await message.answer(f"📜 Умови:\n1. Підписка на {CHANNEL_USERNAME}\n2. YouTube: {YOUTUBE_LINK}\n3. Twitch: {TWITCH_LINK}")
    elif text == "🎁 Призи":
        await message.answer("🎁 Призовий фонд: бонуси для 3 учасників!")
    elif text == "📍 Мій статус":
        status = "✅ Ви берете участь!" if user_id in participants_set else "❌ Ви ще не брали участі."
        await message.answer(status)
    elif text == "📞 Підтримка":
        await message.answer("Зв'язок з підтримкою: https://t.me/" + SUPPORT_USERNAME.strip('@'))
    elif text == "❓ FAQ":
        await message.answer("ℹ️ Часті питання:\n- Як дізнатися чи я зареєстрований?\n- Як змінити нікнейм?\n- Як зв'язатися з підтримкою?")
    elif text == "👥 Учасники" and user_id == ADMIN_ID:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        info = "\n".join([f"{row[1]} | {row[2]} | {row[4]} | {row[3]}" for row in ws.iter_rows(min_row=2, values_only=True)])
        await message.answer(f"👥 Список учасників:\n{info}")
    elif text == "📊 Статистика" and user_id == ADMIN_ID:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        count = ws.max_row - 1
        await message.answer(f"📊 Загальна кількість зареєстрованих учасників: {count}")
    elif text == "📥 Експорт Excel" and user_id == ADMIN_ID:
        await message.answer_document(FSInputFile(EXCEL_FILE))
    elif text == "📣 Розсилка" and user_id == ADMIN_ID:
        admin_states[user_id] = "awaiting_broadcast"
        await message.answer("✉️ Введіть повідомлення для розсилки.")
    elif text == "🕒 Планувати розсилку" and user_id == ADMIN_ID:
        admin_states[user_id] = "awaiting_schedule"
        await message.answer("🕒 Введіть дату та час розсилки у форматі YYYY-MM-DD HH:MM та текст:")
    elif text == "⛔ Забанені" and user_id == ADMIN_ID:
        banned_list = "\n".join(map(str, banned_users)) if banned_users else "✅ Немає забанених користувачів."
        await message.answer(f"🚫 Забанені користувачі:\n{banned_list}")
    elif text == "↩️ Повернутись":
        await message.answer("Повертаємося до меню.", reply_markup=user_menu(user_id == ADMIN_ID))
    elif text == "🔐 Admin panel" and user_id == ADMIN_ID:
        await message.answer("🔐 Вхід в адмін-панель.", reply_markup=admin_menu())
    elif user_id in admin_states and admin_states[user_id] == "awaiting_broadcast":
        broadcast_buffer[user_id] = {"text": text.strip()}
        await confirm_broadcast_manual(user_id)
        del admin_states[user_id]
    elif user_id in admin_states and admin_states[user_id] == "awaiting_schedule":
        try:
            parts = text.strip().split(" ", 2)
            date_time = datetime.strptime(f"{parts[0]} {parts[1]}", "%Y-%m-%d %H:%M")
            scheduler.add_job(lambda: asyncio.create_task(bot.send_message(ADMIN_ID, parts[2])), trigger=DateTrigger(run_date=date_time))
            await message.answer(f"🕒 Розсилку заплановано на {date_time}")
        except Exception as e:
            await message.answer("❌ Невірний формат або помилка: " + str(e))
        finally:
            del admin_states[user_id]


async def confirm_broadcast_manual(user_id: int):
    data = broadcast_buffer.pop(user_id, None)
    if not data or "text" not in data:
        await bot.send_message(user_id, "⚠️ Не вдалося отримати текст розсилки. Спробуйте ще раз.")
        return

    message_text = data["text"]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    count = 0
    failed = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            await bot.send_message(row[0], message_text)
            count += 1
        except Exception as e:
            failed.append(row[0])
            logging.warning(f"Не вдалося надіслати {row[0]}: {e}")

    log_wb = load_workbook(LOG_FILE)
    log_ws = log_wb.active
    log_ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        message_text,
        count,
        ", ".join(map(str, failed))
    ])
    log_wb.save(LOG_FILE)

    await bot.send_message(
        user_id,
        f"✅ Розсилку завершено. Повідомлення доставлено {count} учасникам. "
        f"Не вдалося: {len(failed)}"
    )


async def main():
    pool = await asyncpg.create_pool(
        user=DATABASE_USER,
        password=DATABASE_PASSWORD,
        database=DATABASE_NAME,
        host=DATABASE_HOST,
        port=DATABASE_PORT
    )
    dp['db'] = pool
    dp.include_router(router)
    scheduler.start()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
