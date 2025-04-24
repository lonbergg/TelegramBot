import logging
import os
import re
import asyncio
import asyncpg
import pandas as pd
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, Router, types, F
from aiogram.filters import Command
from aiogram.types import (
    InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
    Message, CallbackQuery, FSInputFile
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.date import DateTrigger
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

# Завантажуємо змінні оточення
load_dotenv()

API_TOKEN = os.getenv("API_TOKEN")
# Підтримка кількох адмінів: передаємо через ADMIN_IDS в .env
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()]
DATABASE_HOST = os.getenv("DATABASE_HOST")
DATABASE_PORT = int(os.getenv("DATABASE_PORT", 5432))
DATABASE_USER = os.getenv("DATABASE_USER")
DATABASE_PASSWORD = os.getenv("DATABASE_PASSWORD")
DATABASE_NAME = os.getenv("DATABASE_NAME")
CHANNEL_USERNAME = os.getenv("CHANNEL_USERNAME")
YOUTUBE_LINK = os.getenv("YOUTUBE_LINK")
TWITCH_LINK = os.getenv("TWITCH_LINK")
SUPPORT_USERNAME = os.getenv("SUPPORT_USERNAME")

# Ініціалізація бота, диспетчера та планувальника
bot = Bot(token=API_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler()
router = Router()
logging.basicConfig(level=logging.INFO)

# Шляхи та стани
EXCEL_FILE = 'participants.xlsx'
LOG_FILE = 'broadcast_log.xlsx'
user_states = {}
participants_set = set()
admin_states = {}
banned_users = set()
last_message_times = {}
broadcast_buffer = {}
SPAM_INTERVAL = timedelta(seconds=5)


# Функція для запланованої розсилки
def schedule_broadcast(content: str):
    for aid in ADMIN_IDS:
        asyncio.create_task(bot.send_message(aid, content))


# Функція для сповіщення адміністраторів про помилки
async def notify_admins(text: str):
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(aid, text)
        except Exception:
            pass


# Ініціалізація файлів Excel
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(["Telegram ID", "Username", "Full Name", "Дата участі", "GGPoker Нік", "Email"])
    wb.save(EXCEL_FILE)

if not os.path.exists(LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.append(["Дата", "Повідомлення", "Успішно", "Не вдалося"])
    wb.save(LOG_FILE)


# Функції для роботи з БД та Excel
def export_db_to_excel(pool):
    async def inner():
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM participants")
            df = pd.DataFrame(rows, columns=["telegram_id", "username", "full_name", "joined_at", "nickname", "email"])
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"participants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            df.to_excel(filename, index=False)
            return filename
    return inner


def save_participant(user: types.User, nickname: str, email: str):
    if user.id in participants_set:
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if nickname == row[4] or user.id == row[0]:
            return False
    ws.append([
        user.id,
        f"@{user.username}" if user.username else "(без username)",
        full_name,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        nickname,
        email
    ])
    wb.save(EXCEL_FILE)
    participants_set.add(user.id)
    return True


async def save_participant_to_db(pool, user: types.User, nickname: str, email: str):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO participants (telegram_id, username, full_name, joined_at, nickname, email)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (telegram_id) DO NOTHING
            """,
            user.id,
            f"@{user.username}" if user.username else "(без username)",
            f"{user.first_name or ''} {user.last_name or ''}".strip(),
            datetime.now(),
            nickname,
            email
        )


# ─── ЛОГУВАННЯ РОЗСИЛОК ─────────────────────────────────────────────────────
async def log_broadcast(pool, message_text: str, success_count: int, failed_list: list[int]):
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO broadcast_logs (date, message, success_count, failed_ids)
            VALUES ($1, $2, $3, $4)
            """,
            datetime.now(),
            message_text,
            success_count,
            ", ".join(map(str, failed_list))
        )


# Формування клавіатур
def user_menu(is_admin: bool = False):
    buttons = [
        [KeyboardButton(text="📜 Умови"), KeyboardButton(text="🎁 Призи")],
        [KeyboardButton(text="📞 Підтримка"), KeyboardButton(text="📍 Мій статус")],
        [KeyboardButton(text="🎉 Взяти участь у розігарші"), KeyboardButton(text="❓ FAQ")],
        [KeyboardButton(text="🚫 Поскаржитись")],
    ]
    if is_admin:
        buttons.append([KeyboardButton(text="🔐 Admin panel")])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True, input_field_placeholder="Оберіть опцію")


def support_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✍️ Написати в підтримку")],
            [KeyboardButton(text="🔄 Змінити нікнейм")],
            [KeyboardButton(text="↩️ Назад до меню")],
        ],
        resize_keyboard=True
    )


def admin_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👥 Учасники"), KeyboardButton(text="📥 Експорт Excel")],
            [KeyboardButton(text="📤 Список з бази PostgreSQL")],
            [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="📣 Розсилка")],
            [KeyboardButton(text="🕒 Планувати розсилку"), KeyboardButton(text="⛔ Забанені")],
            [KeyboardButton(text="🗑️ Видалити учасника")],  # нова кнопка
            [KeyboardButton(text="↩️ Повернутись")],
        ],
        resize_keyboard=True
    )


# Хендлер /start
@router.message(Command("start"))
async def welcome_user(message: Message):
    if message.from_user.id in banned_users:
        return
    await message.answer(
        "👋 Вітаємо у GGpoker Telegram боті! Це не просто бот для участі в розіграші, "
        "а також ваш персональний асистент для отримання новин, бонусів та корисної інформації "
        "про GGpoker. Натисніть кнопку нижче, щоб розпочати.",
        reply_markup=user_menu(message.from_user.id in ADMIN_IDS)
    )


# Участь у розігарші
@router.message(F.text == "🎉 Взяти участь у розігарші")
async def participate_command(message: Message):
    if message.from_user.id in banned_users:
        return
    await message.answer("🔄 Обробляємо запит...", disable_notification=True)
    await asyncio.sleep(1.2)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Я підписався", callback_data="participate")]
    ])
    await message.answer(
        f"📋 Для участі в розігарші потрібно:\n"
        f"1. Підписатися на Telegram канал: {CHANNEL_USERNAME}\n"
        f"2. Підписатися на YouTube: {YOUTUBE_LINK}\n"
        f"3. Підписатися на Twitch: {TWITCH_LINK}\n\n"
        "Після цього натисніть кнопку нижче, щоб продовжити.",
        reply_markup=kb
    )


# Перевірка підписки
@router.callback_query(F.data == "participate")
async def check_subscription(callback: CallbackQuery):
    user = callback.from_user
    if user.id in banned_users:
        return
    await callback.message.answer("🔍 Перевіряємо підписку...", disable_notification=True)
    await asyncio.sleep(1.5)
    try:
        member = await bot.get_chat_member(CHANNEL_USERNAME, user.id)
        if member.status not in ["member", "administrator", "creator"]:
            await callback.answer("❌ Спочатку підпишіться на Telegram-канал!", show_alert=True)
            return
    except Exception as e:
        await callback.answer("⚠️ Неможливо перевірити підписку. Спробуйте пізніше.", show_alert=True)
        await notify_admins(f"❗ Помилка перевірки підписки: {e}")
        return

    user_states[user.id] = 'awaiting_nickname'
    await callback.message.answer("✅ Ви приєдналися! Введіть ваш GGPoker нікнейм.")
    await callback.answer()


# Відкриття адмін‑панелі
@router.message(F.text == "🔐 Admin panel")
async def open_admin_panel(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ У вас немає доступу до адмін‑панелі.")
        return
    await message.answer("🔐 Вхід в адмін‑панель.", reply_markup=admin_menu())


# Підтримка
@router.message(F.text == "📞 Підтримка")
async def show_support_options(message: Message):
    await message.answer("Оберіть варіант:", reply_markup=support_menu())


@router.message(F.text == "✍️ Написати в підтримку")
async def handle_write_support(message: Message):
    await message.answer(f"Зв'яжіться з підтримкою тут: https://t.me/{SUPPORT_USERNAME.strip('@')}")


@router.message(F.text == "🔄 Змінити нікнейм")
async def handle_change_nickname(message: Message):
    user_states[message.from_user.id] = "awaiting_new_nickname"
    await message.answer("Введіть новий нікнейм для заміни:")


@router.message(F.text == "🚫 Поскаржитись")
async def handle_complaint(message: Message):
    await message.answer("😔 Якщо у вас є скарга — зверніться до адміністратора.")


@router.message(F.text == "↩️ Назад до меню")
async def back_to_main_menu(message: Message):
    await message.answer("🔙 Повертаємося до головного меню:", reply_markup=user_menu(message.from_user.id in ADMIN_IDS))


# Підтвердження участі
@router.callback_query(F.data == "confirm_participation")
async def confirm_participation(callback: CallbackQuery):
    user_id = callback.from_user.id
    state = user_states.get(user_id)
    if not state or not isinstance(state, dict):
        await callback.answer("⚠️ Щось пішло не так. Спробуйте ще.")
        return
    nickname = state["nickname"]
    email = state["email"]
    if save_participant(callback.from_user, nickname, email):
        await save_participant_to_db(dp['db'], callback.from_user, nickname, email)
        await callback.message.answer(
            "✅ Участь підтверджено! Успіхів!",
            reply_markup=user_menu(user_id in ADMIN_IDS)
        )
    else:
        banned_users.add(user_id)
        await callback.message.answer("🚫 Ви вже брали участь або намагалися обдурити бота.")
    user_states.pop(user_id, None)
    await callback.answer()


# Обробка повідомлень та адмін-розсилки
@router.message()
async def handle_messages(message: Message):
    user_id = message.from_user.id
    text = message.text or ""
    if user_id in banned_users:
        return

    # Антиспам
    now = datetime.now()
    last = last_message_times.get(user_id)
    if last and now - last < SPAM_INTERVAL:
        return
    last_message_times[user_id] = now

    # Реєстрація
    if user_id in user_states:
        state = user_states[user_id]
        if state == 'awaiting_nickname':
            user_states[user_id] = {'step': 'awaiting_email', 'nickname': text.strip()}
            await message.reply("📧 Введіть вашу електронну пошту:")
            return
        elif isinstance(state, dict) and state.get('step') == 'awaiting_email':
            nickname = state['nickname']
            email = text.strip()
            if not re.match(r'^[\w.-]+@[\w.-]+\.\w{2,}$', email):
                await message.reply("❌ Невірний формат email. Спробуйте ще раз:")
                return
            await message.answer(
                "✅ Все готово! Підтвердьте участь:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="✅ Підтверджую участь", callback_data="confirm_participation")]
                ])
            )
            user_states[user_id] = {"nickname": nickname, "email": email}
            return

    # Адмін-команди
    if text == "👥 Учасники" and user_id in ADMIN_IDS:
        rows = await dp['db'].fetch(
            "SELECT telegram_id, username, full_name, nickname, email, joined_at FROM participants"
        )
        if not rows:
            await message.answer("👥 Список учасників порожній.")
        else:
            info = "\n".join(
                f"{r['telegram_id']} | {r['username']} | {r['full_name']} | {r['nickname']} | {r['email']}"
                for r in rows
            )
            await message.answer(f"👥 Список учасників:\n{info}")
        return
    elif text == "🗑️ Видалити учасника" and user_id in ADMIN_IDS:
        admin_states[user_id] = "awaiting_delete"
        await message.answer("🔍 Введіть Telegram ID учасника для видалення:")
        return
    elif admin_states.get(user_id) == "awaiting_delete":
        try:
            pid = int(text.strip())
            await dp['db'].execute("DELETE FROM participants WHERE telegram_id = $1", pid)
            participants_set.discard(pid)
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == pid:
                    ws.delete_rows(idx)
                    break
            wb.save(EXCEL_FILE)
            await message.answer(f"🗑️ Учасника з ID {pid} видалено.")
        except ValueError:
            await message.answer("❌ Неправильний формат ID. Спробуйте ще раз:")
            return
        except Exception as e:
            await message.answer(f"❌ Помилка видалення: {e}")
        finally:
            admin_states.pop(user_id, None)
        return
    elif text == "📤 Список з бази PostgreSQL" and user_id in ADMIN_IDS:
        await message.answer("🔄 Експортуємо список...")
        export_func = export_db_to_excel(dp['db'])
        file_path = await export_func()
        await message.answer_document(FSInputFile(file_path))
        return
    elif text == "📣 Розсилка" and user_id in ADMIN_IDS:
        admin_states[user_id] = "awaiting_broadcast"
        await message.answer("✉️ Введіть текст для розсилки.")
        return
    elif text == "🕒 Планувати розсилку" and user_id in ADMIN_IDS:
        admin_states[user_id] = "awaiting_schedule"
        await message.answer("🕒 Введіть дату, час (YYYY-MM-DD HH:MM) та текст:")
        return
    elif text == "⛔ Забанені" and user_id in ADMIN_IDS:
        banned_list = "\n".join(map(str, banned_users)) or "✅ Список порожній."
        await message.answer(f"🚫 Забанені:\n{banned_list}")
        return
    elif text == "↩️ Повернутись" and user_id in ADMIN_IDS:
        await message.answer("🔙 Повертаємося до головного меню:", reply_markup=user_menu(True))
        return
    # Розсилка
    if user_id in admin_states and admin_states[user_id] == "awaiting_broadcast":
        broadcast_buffer[user_id] = {"text": text.strip()}
        await confirm_broadcast_manual(user_id)
        admin_states.pop(user_id, None)
        return
    if user_id in admin_states and admin_states[user_id] == "awaiting_schedule":
        try:
            date_str, time_str, content = text.split(" ", 2)
            run_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            scheduler.add_job(schedule_broadcast, trigger=DateTrigger(run_date=run_dt), args=[content])
            await message.answer(f"🕒 Розсилку заплановано на {run_dt}")
        except Exception as e:
            await message.answer(f"❌ Помилка: {e}")
        finally:
            admin_states.pop(user_id, None)
        return


# Оновлена розсилка з логуванням і throttle
async def confirm_broadcast_manual(user_id: int):
    data = broadcast_buffer.pop(user_id, None)
    if not data:
        await bot.send_message(user_id, "⚠️ Текст не знайдено.")
        return
    msg = data["text"]
    async with dp['db'].acquire() as conn:
        rows = await conn.fetch("SELECT telegram_id FROM participants")
    count, failed = 0, []
    for r in rows:
        tid = r["telegram_id"]
        try:
            await bot.send_message(tid, msg)
            count += 1
            await asyncio.sleep(0.05)
        except Exception:
            failed.append(tid)
    await log_broadcast(dp['db'], msg, count, failed)
    await bot.send_message(user_id, f"✅ Розсилка: {count} успішно, {len(failed)} помилок.")


# Main
async def main():
    pool = await asyncpg.create_pool(
        user=DATABASE_USER,
        password=DATABASE_PASSWORD,
        database=DATABASE_NAME,
        host=DATABASE_HOST,
        port=DATABASE_PORT
    )
    dp['db'] = pool
    # Ініціалізуємо participants_set із БД для перевірки дублів
    rows = await pool.fetch("SELECT telegram_id FROM participants")
    for r in rows:
        participants_set.add(r['telegram_id'])

    dp.include_router(router)
    scheduler.start()
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

